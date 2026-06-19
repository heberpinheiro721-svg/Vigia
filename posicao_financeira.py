from __future__ import annotations

from pathlib import Path
from datetime import datetime
from io import BytesIO


def listar_posicoes(data_dir: Path) -> list:
    pos_dir = data_dir / 'Posição Financeira'
    if not pos_dir.exists():
        return []
    return sorted(pos_dir.glob('*.xlsx'), key=lambda x: x.name, reverse=True)


def _num(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def parse_posicao(filepath: Path) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # ── QUADRO ────────────────────────────────────────────────────────────────
    ws = wb['QUADRO']
    rows = list(ws.iter_rows(values_only=True))

    def bloco(r_bancos, r_aplic, r_total, c0, c1, c2, c3, c4):
        def parse(r):
            return {
                'valor':       _num(rows[r][c0]),
                'semanal':     _num(rows[r][c1]),
                'mensal':      _num(rows[r][c2]),
                'trimestral':  _num(rows[r][c3]),
                '12m':         _num(rows[r][c4]),
            }
        return {
            'bancos':     parse(r_bancos),
            'aplicacoes': parse(r_aplic),
            'total':      parse(r_total),
        }

    # IAJA: L8=idx7, L9=idx8, L11=idx10  — cols D,E,F,G,H = 3,4,5,6,7
    iaja         = bloco(7,  8,  10, 3, 4, 5, 6, 7)
    # PPG:  mesmas linhas — cols L,M,N,O,P = 11,12,13,14,15
    ppg          = bloco(7,  8,  10, 11, 12, 13, 14, 15)
    # ASSISTENCIAL: L17=idx16, L18=idx17, L20=idx19 — cols D-H
    assistencial = bloco(16, 17, 19, 3, 4, 5, 6, 7)
    # CONSOLIDADO: mesmas linhas — cols L-P
    consolidado  = bloco(16, 17, 19, 11, 12, 13, 14, 15)

    data_ref    = str(rows[6][3]) if rows[6][3] else ''
    cotacao_usd = _num(rows[20][11])

    # Data do relatório — L59 (idx 58): col O='Brasília', col P=datetime
    data_relatorio = None
    for r in rows:
        if r[14] is not None and 'Bras' in str(r[14]) and isinstance(r[15], datetime):
            data_relatorio = r[15].date()
            break

    # ── PREENCHIMENTO ─────────────────────────────────────────────────────────
    wp = wb['Preenchimento']
    prows = list(wp.iter_rows(values_only=True))

    # Linha 1 (idx 0): cabeçalho — datas a partir de col D (idx 3)
    datas_raw = [str(v) for v in prows[0][3:] if v is not None]
    n = len(datas_raw)

    def serie(row_idx: int) -> list:
        return [_num(v) for v in prows[row_idx][3:3 + n]]

    # Reversão: Preenchimento vai do mais recente para o mais antigo
    datas = list(reversed(datas_raw))

    def rev(idx): return list(reversed(serie(idx)))

    # IAJA Total = Bancos (L10=idx9) + Aplicações (L18=idx17)
    total_iaja_h = [b + a for b, a in zip(rev(9), rev(17))]

    # ASSISTENCIAL Total = Bancos (L24=idx23) + Aplicações (L33=idx32)
    total_ass_h  = [b + a for b, a in zip(rev(23), rev(32))]

    # PPG Total em US$ = Bancos (L37=idx36) + Carteira (L40=idx39)
    total_ppg_h  = [b + c for b, c in zip(rev(36), rev(39))]

    # BRASIL = IAJA + ASSISTENCIAL (em R$)
    total_brasil_h = [i + a for i, a in zip(total_iaja_h, total_ass_h)]

    # CONSOLIDADO = Brasil + PPG × cotação (usando cotação atual como proxy histórico)
    total_cons_h = [b + p * cotacao_usd for b, p in zip(total_brasil_h, total_ppg_h)]

    return {
        'data_ref':        data_ref,
        'data_relatorio':  data_relatorio,
        'iaja':            iaja,
        'ppg':             ppg,
        'assistencial':    assistencial,
        'consolidado':     consolidado,
        'cotacao_usd':     cotacao_usd,
        'hist_datas':      datas,
        'hist_brasil':     total_brasil_h,   # IAJA + Assistencial (R$)
        'hist_ppg':        total_ppg_h,      # PPG total (US$)
        'hist_consolidado': total_cons_h,    # Brasil + PPG×cotação (R$)
    }


def _chart_png(titulo: str, datas: list, valores: list, cor: str,
               prefixo: str, w_in: float = 5.0, h_in: float = 2.2) -> bytes | None:
    """Gera PNG de alta resolução via matplotlib."""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import matplotlib.ticker as mticker

        validos = [v for v in valores if v and v > 0]
        if not validos:
            return None

        y_min  = min(validos) * 0.97
        y_last = valores[-1]
        y_max  = max(y_min + (y_last - y_min) / 0.80, max(validos) * 1.03)
        n      = len(valores)
        x      = list(range(n))

        fig, ax = plt.subplots(figsize=(w_in, h_in), dpi=300)
        ax.fill_between(x, y_min, valores, alpha=0.14, color=cor)
        ax.plot(x, valores, color=cor, linewidth=2.0)
        ax.set_xlim(0, n - 1)
        ax.set_ylim(y_min, y_max)

        # Máximo 10 labels no eixo X para não sobrepor
        step = max(1, n // 10)
        ax.set_xticks(x[::step])
        ax.set_xticklabels(datas[::step], rotation=45, ha='right',
                           fontsize=8, color='#444444')

        def _fmt(v, _):
            if abs(v) >= 1e9: return f'{prefixo} {v/1e9:.2f}Bi'
            if abs(v) >= 1e6: return f'{prefixo} {v/1e6:.0f}Mi'
            return f'{prefixo} {v:,.0f}'

        ax.yaxis.set_major_formatter(mticker.FuncFormatter(_fmt))
        ax.tick_params(axis='y', labelsize=8, colors='#444444')
        ax.set_title(titulo, fontsize=10, color='#1A2E46',
                     fontweight='bold', pad=6)
        ax.set_facecolor('#FFFFFF')
        fig.patch.set_facecolor('#FFFFFF')
        ax.grid(False)
        for sp in ['top', 'right']:
            ax.spines[sp].set_visible(False)
        for sp in ['bottom', 'left']:
            ax.spines[sp].set_color('#DDDDDD')
            ax.spines[sp].set_linewidth(0.8)
        fig.tight_layout(pad=0.6)

        buf = BytesIO()
        plt.savefig(buf, format='png', dpi=300, bbox_inches='tight',
                    facecolor='white', edgecolor='none')
        plt.close(fig)
        buf.seek(0)
        return buf.read()
    except Exception:
        return None


def gerar_pdf_posicao(dados: dict) -> bytes:
    """Gera PDF da Posição Financeira com cabeçalho VIGIA."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable, KeepInFrame,
                                    Image as RLImage)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        topMargin=1.2 * cm, bottomMargin=1.2 * cm,
        leftMargin=1.8 * cm, rightMargin=1.8 * cm,
    )

    COR_AZUL   = colors.HexColor('#1B3A6B')
    COR_LIGHT  = colors.HexColor('#EBF5FB')
    COR_CINZA  = colors.HexColor('#F5F7FA')
    COR_VERDE  = colors.HexColor('#1A7A40')
    COR_VERM   = colors.HexColor('#C0392B')
    BRANCO     = colors.white

    s_vigia = ParagraphStyle('vigia', fontSize=20, textColor=BRANCO,
                              fontName='Helvetica-Bold', alignment=TA_LEFT)
    s_tit   = ParagraphStyle('tit', fontSize=10, textColor=BRANCO,
                              fontName='Helvetica-Bold', alignment=TA_CENTER,
                              leading=14)
    s_nota  = ParagraphStyle('nota', fontSize=7, textColor=colors.grey,
                              fontName='Helvetica', alignment=TA_RIGHT)
    s_secao = ParagraphStyle('secao', fontSize=9, textColor=COR_AZUL,
                              fontName='Helvetica-Bold', alignment=TA_LEFT)
    s_rod   = ParagraphStyle('rod', fontSize=7, textColor=colors.grey,
                              fontName='Helvetica', alignment=TA_CENTER)

    story = []

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    data_str = (dados['data_relatorio'].strftime('%d/%m/%Y')
                if dados['data_relatorio'] else '')
    cab = Table([[
        Paragraph('VIGIA', s_vigia),
        Paragraph(
            'GENERAL CONFERENCE — SOUTH AMERICAN DIVISION<br/>'
            '<b>IAJA / PPG / ASSISTENCIAL</b>',
            s_tit,
        ),
        Paragraph(
            f'Posição: {dados["data_ref"]}<br/>'
            f'Data: {data_str}',
            s_nota,
        ),
    ]], colWidths=['18%', '64%', '18%'])
    cab.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), COR_AZUL),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING',    (0, 0), (-1, -1), 10),
    ]))
    story.append(cab)
    story.append(Spacer(1, 0.3 * cm))

    # ── Helpers ────────────────────────────────────────────────────────────────
    def fmt_val(v, moeda='R$'):
        if abs(v) >= 1e9:
            return f'{moeda} {v / 1e9:.3f} Bi'
        if abs(v) >= 1e6:
            return f'{moeda} {v / 1e6:.2f} Mi'
        return f'{moeda} {v:,.2f}'

    def fmt_pct(v):
        if v == 0:
            return '—'
        p = v * 100
        if abs(p) > 9999:
            return f'{p:+.0f}%'
        return f'{p:+.2f}%'

    def cor_pct(v):
        if v > 0:
            return COR_VERDE
        if v < 0:
            return COR_VERM
        return colors.grey

    def tabela_bloco(titulo, bloco, moeda='R$'):
        cab_row = ['', 'Posição', 'Semanal', 'Mensal', 'Trimestral', '12 Meses']
        linhas = [
            ('Bancos',     bloco['bancos']),
            ('Aplicações', bloco['aplicacoes']),
            ('Total',      bloco['total']),
        ]
        data = [cab_row]
        for nome, d in linhas:
            data.append([
                nome,
                fmt_val(d['valor'], moeda),
                fmt_pct(d['semanal']),
                fmt_pct(d['mensal']),
                fmt_pct(d['trimestral']),
                fmt_pct(d['12m']),
            ])

        cw = [2.3 * cm, 3.2 * cm, 1.8 * cm, 1.8 * cm, 2.2 * cm, 2.2 * cm]
        t = Table(data, colWidths=cw)

        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), COR_AZUL),
            ('TEXTCOLOR',  (0, 0), (-1, 0), BRANCO),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, -1), 7.5),
            ('ALIGN',      (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN',      (0, 0), (0, -1), 'LEFT'),
            ('BACKGROUND', (0, 3), (-1, 3), COR_LIGHT),
            ('FONTNAME',   (0, 3), (-1, 3), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0, 1), (-1, 2), [BRANCO, COR_CINZA]),
            ('GRID',  (0, 0), (-1, -1), 0.3, colors.HexColor('#D0D7E3')),
            ('PADDING', (0, 0), (-1, -1), 4),
        ]
        for ri, (_, d) in enumerate(linhas, start=1):
            for ci, key in enumerate(['semanal', 'mensal', 'trimestral', '12m'], start=2):
                v = d[key]
                if v != 0:
                    style_cmds += [
                        ('TEXTCOLOR', (ci, ri), (ci, ri), cor_pct(v)),
                        ('FONTNAME',  (ci, ri), (ci, ri), 'Helvetica-Bold'),
                    ]
        t.setStyle(TableStyle(style_cmds))

        return [Paragraph(titulo, s_secao), Spacer(1, 0.1 * cm), t]

    # ── 4 tabelas em 2×2 ──────────────────────────────────────────────────────
    w = 13.5 * cm
    b_iaja  = tabela_bloco(f'IAJA — Variação em Reais · {dados["data_ref"]}', dados['iaja'])
    b_ppg   = tabela_bloco(f'PPG — Variação em Dólares · US$ {dados["cotacao_usd"]:.2f}', dados['ppg'], 'US$')
    b_ass   = tabela_bloco('ASSISTENCIAL — Variação em Reais', dados['assistencial'])
    b_cons  = tabela_bloco('CONSOLIDADO — Variação em Reais', dados['consolidado'])

    linha1 = Table([[KeepInFrame(w, 15*cm, b_iaja), KeepInFrame(w, 15*cm, b_ppg)]],
                   colWidths=[w + 0.3*cm, w + 0.3*cm])
    linha2 = Table([[KeepInFrame(w, 15*cm, b_ass), KeepInFrame(w, 15*cm, b_cons)]],
                   colWidths=[w + 0.3*cm, w + 0.3*cm])

    story += [linha1, Spacer(1, 0.25*cm), linha2]

    # ── Cotação US$ em destaque ───────────────────────────────────────────────
    story.append(Spacer(1, 0.2 * cm))
    s_cot_label = ParagraphStyle('cotl', fontSize=7, textColor=BRANCO,
                                 fontName='Helvetica', alignment=TA_RIGHT)
    s_cot_val   = ParagraphStyle('cotv', fontSize=11, textColor=BRANCO,
                                 fontName='Helvetica-Bold', alignment=TA_RIGHT)
    cot_tab = Table([[
        Paragraph('Evolução Histórica', ParagraphStyle(
            'eh', fontSize=9, textColor=COR_AZUL,
            fontName='Helvetica-Bold', alignment=TA_LEFT)),
        Table([[
            Paragraph('COTAÇÃO US$', s_cot_label),
            Paragraph(f'R$ {dados["cotacao_usd"]:.2f}', s_cot_val),
        ]], colWidths=[2.2*cm, 2.0*cm], rowHeights=[0.6*cm],
            style=TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), COR_AZUL),
                ('ALIGN',      (0,0), (-1,-1), 'RIGHT'),
                ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
                ('PADDING',    (0,0), (-1,-1), 4),
                ('ROUNDEDCORNERS', [4]),
            ])),
    ]], colWidths=['82%', '18%'])
    cot_tab.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(cot_tab)
    story.append(Spacer(1, 0.15 * cm))

    # ── Gráficos históricos ───────────────────────────────────────────────────
    datas  = dados['hist_datas']
    cw2    = 12.8 * cm   # largura de cada coluna (2 colunas)
    ch2    = 3.3  * cm   # altura dos 2 gráficos superiores
    ch3    = 3.3  * cm   # altura do gráfico consolidado

    # Figuras geradas maiores que o destino no PDF → texto mais nítido
    png_brasil = _chart_png('Brasil (IAJA + Assistencial)', datas,
                            dados['hist_brasil'], '#1B3A6B', 'R$',
                            w_in=8.0, h_in=2.4)
    png_ppg    = _chart_png('PPG', datas,
                            dados['hist_ppg'], '#27AE60', 'US$',
                            w_in=8.0, h_in=2.4)
    png_cons   = _chart_png(
        f'IAJA Consolidado em R$ (cotação US$ {dados["cotacao_usd"]:.2f})',
        datas, dados['hist_consolidado'], '#2472B5', 'R$',
        w_in=16.0, h_in=2.4)

    def _img(png, w, h):
        return RLImage(BytesIO(png), width=w, height=h) if png else Spacer(w, h)

    linha_g1 = Table(
        [[_img(png_brasil, cw2, ch2), _img(png_ppg, cw2, ch2)]],
        colWidths=[cw2 + 0.2*cm, cw2 + 0.2*cm],
    )
    story.append(linha_g1)
    story.append(Spacer(1, 0.15 * cm))

    if png_cons:
        story.append(RLImage(BytesIO(png_cons), width=26.0*cm, height=ch3))

    # ── Rodapé ────────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.25 * cm))
    story.append(HRFlowable(width='100%', thickness=0.5, color=COR_AZUL))
    story.append(Spacer(1, 0.15 * cm))
    story.append(Paragraph(
        f'Brasília, {data_str}  ·  Tesouraria DSA  ·  '
        f'Gerado pelo <b>VIGIA</b> — Sistema de Análise de Investimentos IAJA',
        s_rod,
    ))

    doc.build(story)
    return buf.getvalue()
